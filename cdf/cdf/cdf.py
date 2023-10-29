from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

excel = 'Prediction7.xlsx'
wb = load_workbook(excel, data_only=True)
ws = wb.active
range = ws['CY2':'CY1501']
data = []
for cell in range:
    for c in cell:
        data.append(c.value)
adata = np.array(data)
x = np.sort(adata)
y = np.arange(len(x)) / float(len(x))
plt.xlabel('Distance')
plt.ylabel('CDF')
plt.title('CDF using distance in ' + excel)
plt.plot(x, y, marker='o')
plt.show()